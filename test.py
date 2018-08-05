import os
import uuid
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def paddingAmount(amount):
    return "{:10.2f}".format(amount).replace('.', '')


def daxie(n):
    if n == ' ':
        return ' '

    return ' ' + '零壹贰叁肆伍陆柒捌玖'[int(n)] + ' '


daxieUnit = ['佰', '拾', '万', '仟', '佰', '拾', '元', '角', '分']


items = 5
itemTotal = paddingAmount(1024)
total = paddingAmount(1024 * items)
path = ''

rows = [3, 5, 7, 9, 11, 13, 15]
for row in rows:
    if items <= row:
        path = 'cost-{}.xlsx'.format(row)
        break

cwd = os.getcwd()
wb = load_workbook(cwd + '/xlsx-templates/{}'.format(path))
ws = wb.active

## 报销项目
for i in range(items):
    r = str(i + 5)
    ws['B' + r] = 'hello'
    ws['B' + r].alignment = Alignment(horizontal='center', vertical='center')
    ws['C' + r] = 'world'
    for j, ch in enumerate(itemTotal):
        col = chr(ord('E') + j)
        ws[col + r] = ch

## 合计
r = 5 + row
for j, ch in enumerate(total):
    col = chr(ord('E') + j)
    ws[col + str(r)] = ch

## 大写金额
r = r + 1
daxieText = '金额大写：'
first = True
for index, ch in enumerate(total):
    dx = daxie(ch)
    if first and dx == ' ':
        continue

    first = False
    daxieText = daxieText + daxie(ch) + daxieUnit[index]

ws['B' + str(r)] = daxieText

## 原借款/退补款
ws['D' + str(r)] = '原借款：{} 元'.format(str(10))
ws['D' + str(r)].alignment = Alignment(horizontal='left', vertical='center')
ws['N' + str(r)] = '退（补）款：{} 元'.format(str(10))

# 报销人/部分负责人/财务负责人
ws['C' + str(r + 1)] = '报销人：hello'
ws['C' + str(r + 2)] = '部门负责人：hello'
ws['C' + str(r + 3)] = '财务负责人：hello'

## 财务会计/人力行政负责人/公司负责人
ws['D' + str(r + 1)] = '财务会计：hello'
ws['D' + str(r + 2)] = '人力行政负责人：hello'
ws['D' + str(r + 3)] = '公司负责人：hello'

## 表头信息
ws['B2'] = '报销部门：{}                {}                      单据及附件共 {} 页'.format('xxxx', 'xxxx', 'xxxx')

## 账号信息系
ws['O3'] = '户名：xxxx\n收款账号:xxxxx \n开户行：xxxxx'
ws['O3'].alignment = Alignment(vertical='center', wrapText=True)

## fix border tyle
thin = Side(border_style="thin", color="000000")
medium = Side(border_style="medium", color="000000")
style_range(ws, 'B3:B4', Border(top=medium, left=medium, right=thin, bottom=thin))
style_range(ws, 'C3:C4', Border(top=medium, left=thin, right=thin, bottom=thin))
style_range(ws, 'E3:M3', Border(top=medium, left=thin, right=thin, bottom=thin))

style_range(ws, 'N3:N' + str(5 + row), Border(top=medium, left=thin, right=thin, bottom=thin))
style_range(ws, 'O3:O' + str(5 + row), Border(top=medium, left=thin, right=medium, bottom=thin))

style_range(ws, 'D{}:M{}'.format(r, r), Border(top=thin, left=thin, right=thin, bottom=thin))
style_range(ws, 'N{}:O{}'.format(r, r), Border(top=thin, left=thin, right=medium, bottom=thin))

style_range(ws, 'D{}:O{}'.format(r + 1, r + 1), Border(top=thin, left=thin, right=medium, bottom=thin))
style_range(ws, 'D{}:O{}'.format(r + 2, r + 2), Border(top=thin, left=thin, right=medium, bottom=thin))
style_range(ws, 'D{}:O{}'.format(r + 3, r + 3), Border(top=thin, left=thin, right=medium, bottom=medium))

style_range(ws, 'B{}:B{}'.format(r + 1, r + 3), Border(top=thin, left=medium, right=thin, bottom=medium))

wb.save('/Users/yangchen/Downloads/export_cost{}.xlsx'.format(str(uuid.uuid4())))
