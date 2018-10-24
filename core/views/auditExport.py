from __future__ import unicode_literals
from decimal import Decimal
import os
import re
import json
import logging
import datetime
import zipfile
from shutil import copyfile, make_archive

import iso8601
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.conf import settings
from sendfile import sendfile
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

from core.models import *
from core.auth import validateToken
from core.common import *

logger = logging.getLogger('app.core.views.auditExport')
thin = Side(border_style="thin", color="000000")
empty = Side(border_style=None, color=None)
medium = Side(border_style="medium", color="000000")


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def amountFixed(amount):
    return "{0:,.2f}".format(amount)


def paddingAmount(amount):
    return "{:10.2f}".format(amount).replace('.', '')


def daxie(n):
    if n == ' ':
        return ' '

    return ' ' + '零壹贰叁肆伍陆柒捌玖'[int(n)] + ' '


daxieUnit = ['佰', '拾', '万', '仟', '佰', '拾', '元', '角', '分']


def convertToDaxieAmountV2(amount):
    pass


def convertToDaxieAmount(amount):
    return convertToDaxieAmountV2(amount)


def convertToDaxieAmountV2(n):
    units = ['', '万', '亿']
    nums = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    decimal_label = ['角', '分']
    small_int_label = ['', '拾', '佰', '仟']
    int_part, decimal_part = str(int(n)), str(n - int(n))[2:]  # 分离整数和小数部分

    res = []
    if decimal_part:
        res.append(''.join([nums[int(x)] + y for x, y in list(zip(decimal_part, decimal_label)) if x != '0']))

    if int_part != '0':
        res.append('元')
        while int_part:
            small_int_part, int_part = int_part[-4:], int_part[:-4]
            tmp = ''.join([nums[int(x)] + (y if x != '0' else '') for x, y in
                           list(zip(small_int_part[::-1], small_int_label))[::-1]])
            tmp = tmp.rstrip('零').replace('零零零', '零').replace('零零', '零')
            unit = units.pop(0)
            if tmp:
                tmp += unit
                res.append(tmp)
    r = ''.join(res[::-1])

    if not decimal_part or decimal_part == '0':
        r = r + '整'

    return r


def convertToDaxieAmountV1(amount):
    first = True
    text = ''

    for index, ch in enumerate(paddingAmount(amount)):
        dx = daxie(ch)
        if first and dx == ' ':
            continue

        first = False
        text = text + dx + daxieUnit[index]

    return text


def inBounds(range, cell):
    start, end = range.split(':')
    start_col = ord(start[0]) - ord('A') + 1
    start_row = int(start[1:])
    end_col = ord(end[0]) - ord('A') + 1
    end_row = int(end[1:])
    bounds = [start_col, start_row, end_col, end_row]

    return bounds[0] <= cell.bounds[0] and bounds[1] <= cell.bounds[1] \
           and bounds[2] >= cell.bounds[2] and bounds[3] >= cell.bounds[3]


def resolveDepOwnerStepFromAudit(activity):
    steps = activity.steps()

    for step in steps:
        if step.assigneeDepartment == activity.creator.department and \
                        step.assigneePosition.code == 'owner':
            return step

    # step not found
    return None


def resolveStepFromAudit(activity, dep=None, pos=None):
    steps = activity.steps()

    for step in steps:
        assignee_dep = None
        if step.assigneeDepartment is not None:
            assignee_dep = step.assigneeDepartment.code

        assignee_pos = step.assigneePosition.code

        if assignee_dep == dep and assignee_pos == pos:
            return step

    # step not found
    return None


def resolveProfileFromAudit(activity, dep=None, pos=None):
    steps = activity.steps()

    for step in steps:
        assignee_dep = None
        if step.assigneeDepartment is not None:
            assignee_dep = step.assigneeDepartment.code

        assignee_pos = step.assigneePosition.code

        if assignee_dep == dep and assignee_pos == pos:
            return step.assignee

    # profile not found
    return None


def fix_merged_cells_border(ws, range):
    for cell in ws.merged_cells:
        if not inBounds(range, cell):
            continue

        style_range(ws, cell.coord, Border(top=thin, left=thin, right=thin, bottom=thin))


def exportOpenAccountAuditDoc(activity):
    account = activity.extra['account']

    cwd = os.getcwd()
    wb = load_workbook(cwd + '/xlsx-templates/open_account.xlsx')
    ws = wb.active

    ws['B2'].value = activity.creator.department.name
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

    ws['D2'].value = activity.created_at.strftime('%Y-%m-%d')
    ws['D2'].alignment = Alignment(horizontal='center', vertical='center')

    ws['B3'].value = account['name']
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center')

    ws['B4'].value = account['bank']
    ws['B4'].alignment = Alignment(horizontal='center', vertical='center')
    natures = [
        {'value': "basic", 'label': "基本账户"},
        {'value': "normal", 'label': "一般账户"},
        {'value': "temporary", 'label': "临时账户"},
        {'value': "special", 'label': "专用账户"}
    ]
    nature = account['nature']
    for n in natures:
        if n['value'] == nature:
            ws['D4'].value = n['label']
            ws['D4'].alignment = Alignment(horizontal='center', vertical='center')

    ws['B5'].value = account['reason']
    ws['B5'].alignment = Alignment(horizontal='center', vertical='center')

    ws['B6'].value = activity.creator.name
    ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
    owner = activity.creator.owner
    ws['D6'].value = owner.name if owner is not None else ''
    ws['D6'].alignment = Alignment(horizontal='center', vertical='center')

    finOwner = resolveProfileFromAudit(activity, 'fin', 'onwer')
    # finOwner = Profile.objects.filter(department__code='fin', position__code='owner', archived=False).first()
    ws['B7'].value = finOwner.name if finOwner is not None else ''
    ws['B7'].alignment = Alignment(horizontal='center', vertical='center')

    ceo = resolveProfileFromAudit(activity, dep='root', pos='ceo')
    # ceo = Profile.objects.filter(position__code='ceo', archived=False).first()
    ws['D7'].value = ceo.name if ceo is not None else ''
    ws['D7'].alignment = Alignment(horizontal='center', vertical='center')

    ws['B8'].value = activity.created_at.strftime('%Y-%m-%d')
    ws['B8'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B9'].value = account.get('desc', '')
    ws['B9'].alignment = Alignment(horizontal='center', vertical='center')

    # fix border style
    for cell in ws.merged_cells:
        if not inBounds('A2:D9', cell):
            logger.info('ignore cell: {}'.format(cell.coord))
            continue

        logger.info('fix border style {}'.format(cell.coord))
        style_range(ws, cell.coord, Border(top=thin, left=thin, right=thin, bottom=thin))

    set_border(ws, 'A2:D9', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportCostAuditDoc(activity):
    items = activity.extra['items']
    template = ''
    rows = [3, 5, 7, 9, 11, 13, 15]
    for row in rows:
        if len(items) <= row:
            template = 'cost-{}.xlsx'.format(row)
            break

    wb = load_workbook(os.getcwd() + '/xlsx-templates/{}'.format(template))
    ws = wb.active

    # cost items
    first_item_row = 5
    totalAmount = 0
    tuibukuan = Decimal('0')
    yuanjiekuan = Decimal('0')
    for i in range(len(items)):
        item = items[i]

        r = str(first_item_row + i)
        ws['B' + r] = item['name']
        ws['B' + r].alignment = Alignment(horizontal='center', vertical='center')
        ws['C' + r] = item['desc']
        ws['C' + r].alignment = Alignment(horizontal='center', vertical='center')
        ws['B' + r].alignment = Alignment(horizontal='center', vertical='center')
        tuibukuan = tuibukuan + Decimal(item['tuibukuan'])
        yuanjiekuan = yuanjiekuan + Decimal(item['yuanjiekuan'])

        amount = float(item['amount'])
        totalAmount = amount + totalAmount

        amount = paddingAmount(float(item['amount']))
        for j, ch in enumerate(amount):
            col = chr(ord('E') + j)
            ws[col + r] = ch
            ws[col + r].alignment = Alignment(horizontal='center', vertical='center')

    # 统计
    r = 5 + row
    for j, ch in enumerate(paddingAmount(totalAmount)):
        col = chr(ord('E') + j)
        ws[col + str(r)] = ch
        ws[col + str(r)].alignment = Alignment(horizontal='center', vertical='center')

    # 大写金额
    r = r + 1
    daxieText = '金额大写：' + convertToDaxieAmount(totalAmount)
    ws['B' + str(r)] = daxieText

    ws['D' + str(r)] = '原借款：{} 元'.format(amountFixed(yuanjiekuan))
    ws['D' + str(r)].alignment = Alignment(horizontal='left', vertical='center')
    ws['N' + str(r)] = '退（补）款：{} 元'.format(amountFixed(tuibukuan))
    ws['N' + str(r)].alignment = Alignment(horizontal='left', vertical='center')

    creator = activity.creator
    owner = creator.owner
    finOwner = resolveProfileFromAudit(activity, dep='fin', pos='owner')
    finAccountant = resolveProfileFromAudit(activity, dep='fin', pos='fin_accountant')
    hrOwner = resolveProfileFromAudit(activity, dep='hr', pos='owner')
    # ceo = Profile.objects.filter(department__code='root', position__code='ceo', archived=False).first()
    ceo = resolveProfileFromAudit(activity, dep='root', pos='ceo')

    # 报销人/部分负责人/财务负责人
    ws['C' + str(r + 1)] = '报销人：{}'.format(getattr(creator, 'name', ''))
    ws['C' + str(r + 2)] = '部门负责人：{}'.format(getattr(owner, 'name', ''))
    ws['C' + str(r + 3)] = '财务负责人：{}'.format(getattr(finOwner, 'name'))

    ## 财务会计/人力行政负责人/公司负责人
    ws['D' + str(r + 1)] = '财务会计：{}'.format(getattr(finAccountant, 'name', ''))
    ws['D' + str(r + 2)] = '人力行政负责人：{}'.format(getattr(hrOwner, 'name', ''))
    ws['D' + str(r + 3)] = '公司负责人：{}'.format(getattr(ceo, 'name', ''))

    ## 表头信息
    dp = creator.department.name if creator.department is not None else ''
    date = datetime.datetime.now().strftime('%Y-%m-%d')
    ws['B2'] = '报销部门：{}'.format(dp)
    ws['E2'] = date
    ws['L2'] = '单据及附件共 {} 页'.format(str(len(items)))

    ## 账号信息系
    account = activity.extra['account']
    ws['O3'] = '户名：{}\n收款账号:{} \n开户行：{}' \
        .format(account['name'], account['number'], account['bank'])
    ws['O3'].alignment = Alignment(vertical='center', wrapText=True)

    # fix border tyle
    style_range(ws, 'B3:B4', Border(top=medium, left=medium, right=thin, bottom=thin))
    style_range(ws, 'C3:C4', Border(top=medium, left=thin, right=thin, bottom=thin))
    style_range(ws, 'E3:M3', Border(top=medium, left=thin, right=thin, bottom=thin))

    style_range(ws, 'N3:N' + str(5 + row), Border(top=medium, left=thin, right=thin, bottom=thin))
    style_range(ws, 'O3:O' + str(5 + row), Border(top=medium, left=thin, right=medium, bottom=thin))

    style_range(ws, 'D{}:M{}'.format(r, r), Border(top=thin, left=thin, right=thin, bottom=thin))
    style_range(ws, 'N{}:O{}'.format(r, r), Border(top=thin, left=thin, right=medium, bottom=thin))

    style_range(ws, 'D{}:O{}'.format(r + 1, r + 1), Border(top=thin, left=thin, right=medium, bottom=thin))
    style_range(ws, 'D{}:O{}'.format(r + 2, r + 2), Border(top=thin, left=thin, right=medium, bottom=thin))
    style_range(ws, 'D{}:O{}'.format(r + 3, r + 3), Border(top=thin, left=thin, right=medium, bottom=medium))

    style_range(ws, 'B{}:B{}'.format(r + 1, r + 3), Border(top=thin, left=medium, right=thin, bottom=medium))

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportLoanAuditDoc(activity):
    wb = load_workbook(os.getcwd() + '/xlsx-templates/loan.xlsx')
    ws = wb.active

    auditData = activity.extra
    creator = activity.creator
    ws['B3'] = '部门:{}                {}                      编号:{}'. \
        format(getattr(creator.department, 'name', ''),
               datetime.datetime.now().strftime('%Y-%m-%d'),
               activity.sn)

    # 人民币大写
    amount = float(auditData['loan']['amount'])
    ws['B5'] = '人民币（大写） {}  此据'.format(convertToDaxieAmount(amount))
    ws['B5'].alignment = Alignment(vertical='center', wrapText=True)

    # 人民币小写
    ws['B6'] = '（小写）￥ {}'.format(amountFixed(float(auditData['loan']['amount'])))
    ws['B6'].alignment = Alignment(vertical='center', wrapText=True)

    # 借款用途说明
    ws['C7'] = auditData['loan']['application']
    ws['C7'].alignment = Alignment(vertical='center', wrapText=True)

    # 付款信息
    account = auditData['account']
    ws['C8'] = '户名：{}\n收款账号:{} \n开户行：{}' \
        .format(account['name'], account['number'], account['bank'])
    ws['C8'].alignment = Alignment(vertical='center', wrapText=True)

    # 借款人、部门负责人、财务负责人、公司负责人
    ws['C11'] = creator.name
    ws['C11'].alignment = Alignment(vertical='center', horizontal='center')

    ws['F11'] = getattr(creator.owner, 'name', '')
    ws['F11'].alignment = Alignment(vertical='center', horizontal='center')

    # finOwner = Profile.objects.filter(department__code='fin', position__code='owner', archived=False).first()
    finOwner = resolveProfileFromAudit(activity, dep='fin', pos='owner')
    ws['I11'] = getattr(finOwner, 'name', '')
    ws['I11'].alignment = Alignment(vertical='center', horizontal='center')

    # ceo = Profile.objects.filter(department__code='root', position__code='ceo', archived=False).first()
    ceo = resolveProfileFromAudit(activity, dep='root', pos='ceo')
    ws['M11'] = getattr(ceo, 'name', '')
    ws['M11'].alignment = Alignment(vertical='center', horizontal='center')

    # fix border style
    style_range(ws, 'B4:P4', Border(top=medium, left=medium, right=medium, bottom=thin))
    style_range(ws, 'B5:P5', Border(top=thin, left=medium, right=medium, bottom=thin))
    style_range(ws, 'B6:P6', Border(top=thin, left=medium, right=medium, bottom=thin))
    style_range(ws, 'C7:P7', Border(top=thin, left=thin, right=medium, bottom=thin))
    style_range(ws, 'B8:B9', Border(top=thin, left=medium, right=thin, bottom=thin))
    style_range(ws, 'C8:P9', Border(top=thin, left=thin, right=medium, bottom=thin))

    style_range(ws, 'B10:B11', Border(top=thin, left=medium, right=thin, bottom=medium))

    style_range(ws, 'C10:E10', Border(top=thin, left=thin, right=thin, bottom=thin))
    style_range(ws, 'C11:E11', Border(top=thin, left=thin, right=thin, bottom=medium))
    style_range(ws, 'F10:H10', Border(top=thin, left=thin, right=thin, bottom=thin))
    style_range(ws, 'F11:H11', Border(top=thin, left=thin, right=thin, bottom=medium))
    style_range(ws, 'I10:L10', Border(top=thin, left=thin, right=thin, bottom=thin))
    style_range(ws, 'I11:L11', Border(top=thin, left=thin, right=thin, bottom=medium))
    style_range(ws, 'M10:P10', Border(top=thin, left=thin, right=medium, bottom=thin))
    style_range(ws, 'M11:P11', Border(top=thin, left=thin, right=medium, bottom=medium))

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def set_border(ws, cell_range, border_style):
    rows = ws[cell_range]
    side = Side(border_style=border_style, color="00000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def exportMoneyAuditDoc(activity):
    wb = load_workbook(os.getcwd() + '/xlsx-templates/money.xlsx')
    ws = wb.active

    ws['A1'] = '用 款 申 请 单'
    ws['A1'].alignment = Alignment(vertical='center', horizontal='center')

    auditData = activity.extra
    creator = activity.creator
    ws['A2'] = '  用款部门:{}'.format(getattr(creator.department, 'name', ''))
    ws['K2'] = datetime.datetime.now().strftime('%Y-%m-%d')

    # 收款信息
    info = auditData['info']
    inAccount = auditData['inAccount']
    ws['B3'] = inAccount['name']
    ws['B4'] = inAccount['bank']
    ws['H4'] = inAccount['number']
    ws['B5'] = '（大写）{}'.format(convertToDaxieAmount(float(info['amount'])))
    ws['J5'] = '￥' + amountFixed(float(info['amount']))

    # 出款信息
    outAccount = auditData['outAccount']
    ws['B6'] = outAccount['name']
    ws['B6'].alignment = Alignment(vertical='center', horizontal='center')
    ws['k6'] = '现金' if outAccount['type'] == 'cash' else '转账'
    ws['B7'] = outAccount['bank']
    ws['H7'] = outAccount['number']

    ws['B8'] = info['desc']
    ws['B8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 借款人、部门负责人、财务负责人、公司负责人
    ws['B10'] = creator.name
    ws['B10'].alignment = Alignment(vertical='center', horizontal='center')

    ws['H10'] = getattr(creator.owner, 'name', '')
    ws['H10'].alignment = Alignment(vertical='center', horizontal='center')

    accountant = resolveProfileFromAudit(activity, dep='fin', pos='fin_accountant')
    ws['B12'] = getattr(accountant, 'name', '')
    ws['B12'].alignment = Alignment(vertical='center', horizontal='center')

    finOwner = resolveProfileFromAudit(activity, dep='fin', pos='owner')
    ws['H12'] = getattr(finOwner, 'name', '')
    ws['H12'].alignment = Alignment(vertical='center', horizontal='center')

    ceo = resolveProfileFromAudit(activity, dep='root', pos='ceo')
    ws['H13'] = getattr(ceo, 'name', '')
    ws['H13'].alignment = Alignment(vertical='center', horizontal='center')

    # A3:M3
    for cell in ws.merged_cells:
        if not inBounds('A3:M13', cell):
            logger.info('ignore cell: {}'.format(cell.coord))
            continue

        logger.info('fix border style {}'.format(cell.coord))
        style_range(ws, cell.coord, Border(top=thin, left=thin, right=thin, bottom=thin))
    set_border(ws, 'A3:M13', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportBizContractAuditDoc(activity):
    wb = load_workbook(os.getcwd() + '/xlsx-templates/biz_contract.xlsx')
    ws = wb.active

    creator = activity.creator
    auditData = activity.extra
    base = auditData['base']
    info = auditData['info']

    ws[
        'A3'] = '合同类型：{}                                                                                                              {}' \
        .format('大宗类' if base['type'] == 'dazong' else '其他类', datetime.datetime.now().strftime('%Y-%m-%d'))
    ws['B4'] = base.get('company', '')

    ws['B5'] = info['upstream']
    ws['B5'].alignment = Alignment(vertical='center', horizontal='center')

    ws['F5'] = info.get('downstream', '')
    ws['B6'] = info['asset']
    ws['B7'] = amountFixed(float(info['tonnage'])) + '吨'
    ws['F7'] = amountFixed(float(info['buyPrice'])) + '元/吨'

    ws['B8'] = '现金' if info['settlementType'] == 'cash' else '转账'
    ws['F8'] = amountFixed(float(info['sellPrice'])) + '元/吨'
    ws['B9'] = info['profitsPerTon']
    ws['F9'] = info['grossMargin'] + '%'
    ws['B10'] = info.get('desc', '')
    ws['B10'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B11'] = creator.name
    ws['B11'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F11'] = getattr(creator.owner, 'name', '')
    ws['F11'].alignment = Alignment(horizontal='center', vertical='center')

    accountant = resolveProfileFromAudit(activity, dep='fin', pos='fin_accountant')
    ws['F12'] = getattr(accountant, 'name', '')
    ws['F12'].alignment = Alignment(horizontal='center', vertical='center')

    # TODO: 法务负责人
    finOwner = resolveProfileFromAudit(activity, dep='fin', pos='owner')
    ws['B13'] = getattr(finOwner, 'name', '')
    ws['B13'].alignment = Alignment(horizontal='center', vertical='center')

    ceo = resolveProfileFromAudit(activity, dep='root', pos='ceo')
    ws['B14'] = getattr(ceo, 'name', '')
    ws['B14'].alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws.merged_cells:
        if not inBounds('A4:F19', cell):
            logger.info('ignore cell: {}'.format(cell.coord))
            continue

        logger.info('fix border style {}'.format(cell.coord))
        style_range(ws, cell.coord, Border(top=thin, left=thin, right=thin, bottom=thin))

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportFnContractAuditDoc(activity):
    auditData = activity.extra
    info = auditData['info']
    base = auditData['base']
    template = 'fn_contract'
    if float(info['amount']) == 0:
        template = 'fn_contract_zero'

    wb = load_workbook(os.getcwd() + '/xlsx-templates/{}.xlsx'.format(template))
    ws = wb.active

    ws['B3'] = base['name']
    ws['D3'] = base['other']

    ws['B4'] = float(info['amount'])
    ws['B4'].number_format = '#,##0.00'
    ws['D4'] = info.get('date', '')
    ws['B5'] = info['count']
    ws['D5'] = info['sn']
    ws['B6'] = info.get('desc', '')
    ws['B6'].alignment = Alignment(horizontal='center', vertical='center')

    # FIXME: 法务负责人
    steps = activity.steps()
    for step in steps:
        dp = None
        if step.assigneeDepartment is not None:
            dp = step.assigneeDepartment.code
        pos = step.assigneePosition.code

        desc = step.desc

        if dp is None and pos == 'owner':
            ws['B7'] = step.desc if desc is not None and desc != '' else '同意'

        if template == 'fn_contract':
            if dp == 'fin' and pos == 'owner':
                ws['B9'] = step.desc if step.desc is not None and desc != '' else '同意'
            if dp == 'root' and pos == 'ceo':
                ws['B10'] = step.desc if step.desc is not None and desc != '' else '同意'
        else:
            if dp == 'root' and pos == 'ceo':
                ws['B9'] = step.desc if step.desc is not None and desc != '' else '同意'

    for cell in ws.merged_cells:
        if not inBounds('A3:D11', cell):
            logger.info('ignore cell: {}'.format(cell.coord))
            continue

        logger.info('fix border style {}'.format(cell.coord))
        style_range(ws, cell.coord, Border(top=thin, left=thin, right=thin, bottom=thin))

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportTravelAuditDoc(activity):
    auditData = activity.extra
    creator = activity.creator

    items = auditData['items']
    availableRows = 4
    if len(items) > availableRows:
        availableRows = 10

    if availableRows == 4:
        wb = load_workbook(os.getcwd() + '/xlsx-templates/travel.xlsx')
    else:
        wb = load_workbook(os.getcwd() + '/xlsx-templates/travel-10.xlsx')

    ws = wb.worksheets[0]
    ws['C3'] = '姓名: {}                 部门: {}                    {}'.format(
        creator.name,
        getattr(creator.department, 'name', ''),
        datetime.datetime.now().strftime('%Y-%m-%d')
    )

    firstItemRow = 8
    total = {
        'normal': 0,
        'train': 0,
        'car': 0,
        'ship': 0,
        'plane': 0,
        'hotel': 0,
        'traffic': 0,
        'other': 0,
        'amount2': 0
    }
    t = 0
    for index, item in enumerate(items):
        def parseDate(str):
            return datetime.datetime.strptime(str, '%Y-%m-%d %H:%M:%S')

        startDate, endDate = parseDate(item['startTime']), parseDate(item['endTime'])
        row = str(firstItemRow + index)
        ws['C' + row] = startDate.month
        ws['C' + row].alignment = Alignment(vertical='center', horizontal='center')
        ws['D' + row] = startDate.day
        ws['D' + row].alignment = Alignment(vertical='center', horizontal='center')
        ws['E' + row] = startDate.hour
        ws['E' + row].alignment = Alignment(vertical='center', horizontal='center')
        ws['F' + row] = endDate.month
        ws['F' + row].alignment = Alignment(vertical='center', horizontal='center')
        ws['G' + row] = endDate.day
        ws['G' + row].alignment = Alignment(vertical='center', horizontal='center')
        ws['H' + row] = endDate.hour
        ws['H' + row].alignment = Alignment(vertical='center', horizontal='center')
        days = item['days']
        ws['I' + row] = days
        ws['I' + row].alignment = Alignment(vertical='center', horizontal='center')

        ws['J' + row] = item['place']
        ws['J' + row].alignment = Alignment(vertical='center', horizontal='center')
        ws['K' + row] = days
        ws['K' + row].alignment = Alignment(vertical='center', horizontal='center')
        ws['L' + row] = item['spec']
        ws['L' + row].alignment = Alignment(vertical='center', horizontal='center')

        normal = float(item['spec']) * days
        total['normal'] = total['normal'] + normal
        ws['M' + row] = normal
        ws['M' + row].alignment = Alignment(vertical='center', horizontal='center')

        train = float(item.get('train', '0'))
        total['train'] = total['train'] + train
        ws['N' + row] = amountFixed(train)
        ws['N' + row].alignment = Alignment(vertical='center', horizontal='center')

        car = float(item.get('car', '0'))
        total['car'] = total['car'] + car
        ws['O' + row] = amountFixed(car)
        ws['O' + row].alignment = Alignment(vertical='center', horizontal='center')

        ship = float(item.get('ship', '0'))
        total['ship'] = total['ship'] + ship
        ws['P' + row] = amountFixed(ship)
        ws['P' + row].alignment = Alignment(vertical='center', horizontal='center')

        plane = float(item.get('plane', '0'))
        total['plane'] = total['plane'] + plane
        ws['Q' + row] = amountFixed(plane)
        ws['Q' + row].alignment = Alignment(vertical='center', horizontal='center')

        hotel = float(item.get('hotel', '0'))
        total['hotel'] = total['hotel'] + hotel
        ws['R' + row] = amountFixed(hotel)
        ws['R' + row].alignment = Alignment(vertical='center', horizontal='center')

        traffic = float(item.get('traffic', '0'))
        total['traffic'] = total['traffic'] + traffic
        ws['S' + row] = amountFixed(traffic)
        ws['S' + row].alignment = Alignment(vertical='center', horizontal='center')

        other = float(item.get('other', '0'))
        total['other'] = total['other'] + other
        ws['T' + row] = amountFixed(other)
        ws['T' + row].alignment = Alignment(vertical='center', horizontal='center')

        amount1 = train + car + ship + plane + traffic + other + hotel
        ws['V' + row] = amountFixed(amount1)
        ws['V' + row].alignment = Alignment(vertical='center', horizontal='center')

        amount2 = normal + amount1
        total['amount2'] = amount2 + total['amount2']
        ws['W' + row] = amountFixed(amount2)
        ws['W' + row].alignment = Alignment(vertical='center', horizontal='center')

        t = amount2 + t

    r = firstItemRow + availableRows
    ws['k' + str(r)] = amountFixed(total['normal'])
    ws['N' + str(r)] = amountFixed(total['train'])
    ws['O' + str(r)] = amountFixed(total['car'])
    ws['P' + str(r)] = amountFixed(total['ship'])
    ws['Q' + str(r)] = amountFixed(total['plane'])
    ws['R' + str(r)] = amountFixed(total['hotel'])
    ws['S' + str(r)] = amountFixed(total['traffic'])
    ws['T' + str(r)] = amountFixed(total['other'])
    ws['U' + str(r)] = '0'
    ws['V' + str(r)] = amountFixed(total['amount2'])
    ws['W' + str(r)] = amountFixed(t)

    ws['C' + str(r + 1)] = '合计人民币（大写）    {}'.format(convertToDaxieAmount(t))
    info = auditData['info']
    ws['C' + str(r + 2)] = '原借差旅费 {} 元                 剩余交回 {} 元'.format(
        amountFixed(float(info['yuanjiekuan'])),
        amountFixed(float(info.get('shengyu', '0')))
    )
    ws['D' + str(r + 3)] = info.get('reason', '')
    ws['X4'] = '附件共（{}）张'.format(len(auditData.get('attachments', [])))

    finAccountant = resolveProfileFromAudit(activity, dep='fin', pos='fin_accoutant')
    hr = resolveProfileFromAudit(activity, dep='hr', pos='hr_member')
    ws['C' + str(r + 5)] = '会计：{}              人资专员：{}              出差人员签字：{}'.format(
        getattr(finAccountant, 'name', ''), getattr(hr, 'name', ''), creator.name)

    for cell in ws.merged_cells:
        if not inBounds('C4:W' + str(r + 4), cell):
            logger.info('ignore cell: {}'.format(cell.coord))
            continue

        logger.info('fix border style {}'.format(cell.coord))
        style_range(ws, cell.coord, Border(top=thin, left=thin, right=thin, bottom=thin))

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    # 费用报销
    ws = wb.worksheets[1]
    info = auditData['info']
    dp = creator.department.name if creator.department is not None else ''
    date = datetime.datetime.now().strftime('%Y-%m-%d')
    ws['B2'] = '报销部门：{}'.format(dp)
    ws['E2'] = date
    ws['L2'] = '单据及附件共 1 页'

    ws['B5'] = '差旅报销'
    ws['C5'] = getattr(info, 'reason', '')
    amount = paddingAmount(t)
    for j, ch in enumerate(amount):
        col = chr(ord('E') + j)
        ws[col + '5'] = ch
        ws[col + '5'].alignment = Alignment(vertical='center', wrapText=True)

    ## 合计
    for j, ch in enumerate(amount):
        col = chr(ord('E') + j)
        ws[col + '8'] = ch
        ws[col + '8'].alignment = Alignment(vertical='center', wrapText=True)

    ## 金额大写
    ws['B9'] = '金额大写：{}'.format(convertToDaxieAmount(t))

    ws['D9'] = '原借款：{} 元'.format(amountFixed(float(info['yuanjiekuan'])))
    if info.get('tuibukuan', None) is not None:
        ws['N9'] = '退补款：{} 元'.format(amountFixed(float(info.get('tuibukuan'))))

    creator = activity.creator
    owner = creator.owner
    finOwner = resolveProfileFromAudit(activity, dep='fin', pos='owner')
    finAccountant = resolveProfileFromAudit(activity, dep='fin', pos='fin_accoutant')
    hrOwner = resolveProfileFromAudit(activity, dep='hr', pos='owner')
    ceo = resolveProfileFromAudit(activity, dep='root', pos='ceo')
    # 报销人/部分负责人/财务负责人
    ws['C10'] = '报销人：{}'.format(getattr(creator, 'name', ''))
    ws['C11'] = '部门负责人：{}'.format(getattr(owner, 'name', ''))
    ws['C12'] = '财务负责人：{}'.format(getattr(finOwner, 'name'))

    ## 财务会计/人力行政负责人/公司负责人
    ws['D10'] = '财务会计：{}'.format(getattr(finAccountant, 'name', ''))
    ws['D11'] = '人力行政负责人：{}'.format(getattr(hrOwner, 'name', ''))
    ws['D12'] = '公司负责人：{}'.format(getattr(ceo, 'name', ''))

    ## 账号信息系
    account = auditData['account']
    ws['O3'] = '户名：{}\n收款账号:{} \n开户行：{}' \
        .format(account['name'], account['number'], account['bank'])
    ws['O3'].alignment = Alignment(vertical='center', wrapText=True)

    for cell in ws.merged_cells:
        if not inBounds('A3:O12', cell):
            logger.info('ignore cell: {}'.format(cell.coord))
            continue

        logger.info('fix border style {}'.format(cell.coord))
        style_range(ws, cell.coord, Border(top=thin, left=thin, right=thin, bottom=thin))

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportYongrenAuditDoc(activity):
    auditData = activity.extra
    info = auditData['info']
    usage = auditData['usage']
    template = 'yongren'

    wb = load_workbook(os.getcwd() + '/xlsx-templates/{}.xlsx'.format(template))
    ws = wb.active

    ws['B2'] = activity.creator.department.name
    ws['F2'] = activity.created_at.strftime('%Y-%m-%d')
    ws['B3'] = info['position']
    ws['D3'] = info['num']
    ws['F3'] = info['date']
    ws['B4'] = info['desc']
    ws['B4'].alignment = Alignment(vertical='center', wrapText=True)
    gender = '男' if usage['gender'] == '1' else '女'
    ws['B5'] = '性别：{}        年龄：{}\n学历：{}        专业：{}'.format(gender, usage['age'], usage['education'], usage['specs'])
    ws['B5'].alignment = Alignment(vertical='center', wrapText=True)
    ws['B6'] = info['zige']
    ws['B6'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveDepOwnerStepFromAudit(activity)
    ws['B7'] = getattr(step, 'desc', '无')
    ws['B7'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='hr', pos='owner')
    ws['B8'] = getattr(step, 'desc', '无')
    ws['B8'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='root', pos='ceo')
    ws['B9'] = getattr(step, 'desc', '无')
    ws['B9'].alignment = Alignment(vertical='center', wrapText=True)

    fix_merged_cells_border(ws, 'A3:F9')
    set_border(ws, 'A3:F9', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportQingjiaAuditDoc(activity):
    info = activity.extra

    wb = load_workbook(os.getcwd() + '/xlsx-templates/qingjia.xlsx')
    ws = wb.active

    ws['D2'] = activity.created_at.strftime('%Y-%m-%d')
    ws['B3'] = activity.creator.name
    ws['D3'] = activity.creator.position.name
    typeOptions = [
        {'id': 1, 'name': "事假"},
        {'id': 2, 'name': "病假"},
        {'id': 3, 'name': "工伤假"},
        {'id': 4, 'name': "年休假"},
        {'id': 5, 'name': "婚假"},
        {'id': 6, 'name': "产假"},
        {'id': 7, 'name': "产检假"},
        {'id': 8, 'name': "看护假"},
        {'id': 9, 'name': "哺乳假"},
        {'id': 10, 'name': "丧假"},
        {'id': 11, 'name': "考试假"},
        {'id': 100, 'name': "其他"}
    ]
    for o in typeOptions:
        if info['type'] == o['id']:
            ws['B4'] = o['name']
    ws['B4'].alignment = Alignment(vertical='center', horizontal='left', wrapText=True)
    ws['B5'] = info['reason']
    ws['B5'].alignment = Alignment(vertical='center', wrapText=True)

    startDate = datetime.datetime.strptime(info['date'][0], "%Y-%m-%dT%H:%M:%S.%fZ") + datetime.timedelta(hours=8)
    endDate = datetime.datetime.strptime(info['date'][1], "%Y-%m-%dT%H:%M:%S.%fZ") + datetime.timedelta(hours=8)
    ws['B6'] = startDate.strftime('%Y-%m-%d %H:%M') + '-' + endDate.strftime('%Y-%m-%d %H:%M')
    ws['B6'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveDepOwnerStepFromAudit(activity)
    ws['B7'] = getattr(step, 'desc', '无')
    ws['B7'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='hr', pos='owner')
    ws['B8'] = getattr(step, 'desc', '无')
    ws['B8'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='root', pos='ceo')
    ws['B9'] = getattr(step, 'desc', '无')
    ws['B9'].alignment = Alignment(vertical='center', wrapText=True)

    fix_merged_cells_border(ws, 'A3:D9')
    set_border(ws, 'A3:D9', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportChuchaiAuditDoc(activity):
    info = activity.extra

    wb = load_workbook(os.getcwd() + '/xlsx-templates/chuchai.xlsx')
    ws = wb.active

    ws['D2'] = activity.created_at.strftime('%Y-%m-%d')
    ws['B3'] = activity.creator.name
    ws['D3'] = activity.creator.position.name

    ws['B4'] = info['reason']
    ws['B4'].alignment = Alignment(vertical='center', wrapText=True)

    ws['B5'] = info['startTime'] + '-' + info['endTime']
    ws['B5'].alignment = Alignment(vertical='center', wrapText=True)

    ws['B6'] = info['location']
    ws['B6'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveDepOwnerStepFromAudit(activity)
    ws['B7'] = getattr(step, 'desc', '无')
    ws['B7'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='hr', pos='owner')
    ws['B8'] = getattr(step, 'desc', '无')
    ws['B8'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='root', pos='ceo')
    ws['B9'] = getattr(step, 'desc', '无')
    ws['B9'].alignment = Alignment(vertical='center', wrapText=True)

    fix_merged_cells_border(ws, 'A3:D9')
    set_border(ws, 'A3:D9', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportKaoqinyichangAuditDoc(activity):
    info = activity.extra

    wb = load_workbook(os.getcwd() + '/xlsx-templates/kaoqin_yichang.xlsx')
    ws = wb.active

    ws['D2'] = activity.created_at.strftime('%Y-%m-%d')
    ws['B3'] = activity.creator.name
    ws['D3'] = activity.creator.position.name

    ws['B5'] = info['reason']
    ws['B5'].alignment = Alignment(vertical='center', wrapText=True)

    startDate = datetime.datetime.strptime(info['date'][0], "%Y-%m-%dT%H:%M:%S.%fZ") + datetime.timedelta(hours=8)
    endDate = datetime.datetime.strptime(info['date'][1], "%Y-%m-%dT%H:%M:%S.%fZ") + datetime.timedelta(hours=8)
    ws['B4'] = startDate.strftime('%Y-%m-%d %H:%M') + '-' + endDate.strftime('%Y-%m-%d %H:%M')
    ws['B4'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveDepOwnerStepFromAudit(activity)
    ws['B6'] = getattr(step, 'desc', '无')
    ws['B6'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='hr', pos='owner')
    ws['B7'] = getattr(step, 'desc', '无')
    ws['B7'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='root', pos='ceo')
    ws['B8'] = getattr(step, 'desc', '无')
    ws['B8'].alignment = Alignment(vertical='center', wrapText=True)

    fix_merged_cells_border(ws, 'A3:D8')
    set_border(ws, 'A3:D8', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportZizhishiyongAuditDoc(activity):
    info = activity.extra['info']
    usage = activity.extra['usage']

    wb = load_workbook(os.getcwd() + '/xlsx-templates/zizhi_shiyong.xlsx')
    ws = wb.active

    ws['B2'] = info['company']
    ws['B2'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B3'] = info['type']
    ws['B3'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B4'] = activity.creator.name
    ws['D4'] = activity.creator.department.name
    ws['F4'] = activity.created_at.strftime('%Y-%m-%d')

    ws['B5'] = info['desc']
    ws['B5'].alignment = Alignment(vertical='center', horizontal='left')

    step = resolveDepOwnerStepFromAudit(activity)
    ws['B6'] = getattr(step, 'desc', '无')
    ws['B6'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='hr', pos='owner')
    ws['F6'] = getattr(step, 'desc', '无')
    ws['F6'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='fin', pos='owner')
    ws['B7'] = getattr(step, 'desc', '无')
    ws['B7'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='root', pos='ceo')
    ws['F7'] = getattr(step, 'desc', '无')
    ws['F7'].alignment = Alignment(vertical='center', wrapText=True)

    ws['A8'] = '是否带出：{}'.format('是' if usage['out'] == '1' else '否')
    date = usage['date'] if 'date' in usage else '无'
    member = usage['member'] if 'member' in usage else '无'
    desc = usage['desc'] if 'desc' in usage else '无'
    ws['B8'] = '1.预计归还时间：{}             2. 陪同人员：{}\n3.其他说明事项：{}'.format(date, member, desc)
    ws['B8'].alignment = Alignment(vertical='center', wrapText=True)

    fix_merged_cells_border(ws, 'A2:G9')
    set_border(ws, 'A2:G9', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportYinjiankezhiAuditDoc(activity):
    info = activity.extra

    wb = load_workbook(os.getcwd() + '/xlsx-templates/yinjian_kezhi.xlsx')
    ws = wb.active

    ws['B2'] = info['company']
    ws['B2'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B3'] = info['type']
    ws['B3'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B4'] = info['reason']
    ws['B4'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B5'] = info['department']['name']
    ws['B5'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B5'] = info['department']['name']
    ws['B5'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B6'] = activity.creator.name
    ws['B6'].alignment = Alignment(vertical='center', horizontal='left')

    step = resolveDepOwnerStepFromAudit(activity)
    ws['B7'] = getattr(step, 'desc', '无')
    ws['B7'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='root', pos='ceo')
    ws['B9'] = getattr(step, 'desc', '无')
    ws['B9'].alignment = Alignment(vertical='center', wrapText=True)

    fix_merged_cells_border(ws, 'A2:B9')
    set_border(ws, 'A2:B9', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportDanganjiechuAuditDoc(activity):
    info = activity.extra

    wb = load_workbook(os.getcwd() + '/xlsx-templates/dangan_jiechu.xlsx')
    ws = wb.active

    ws['B2'] = activity.creator.name
    ws['B2'].alignment = Alignment(vertical='center', horizontal='left')

    ws['D2'] = activity.creator.department.name
    ws['D2'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B3'] = info['name']
    ws['B3'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B4'] = info['reason']
    ws['B4'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B5'] = info['date']
    ws['B5'].alignment = Alignment(vertical='center', horizontal='left')

    step = resolveDepOwnerStepFromAudit(activity)
    ws['B6'] = getattr(step, 'desc', '无')
    ws['B6'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='root', pos='ceo')
    ws['B8'] = getattr(step, 'desc', '无')
    ws['B8'].alignment = Alignment(vertical='center', wrapText=True)

    fix_merged_cells_border(ws, 'A2:D9')
    set_border(ws, 'A2:D9', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def exportZichanbaofeiAuditDoc(activity):
    info = activity.extra

    wb = load_workbook(os.getcwd() + '/xlsx-templates/zichan_baofei.xlsx')
    ws = wb.active

    ws['B2'] = activity.creator.department.name
    ws['B2'].alignment = Alignment(vertical='center', horizontal='left')

    ws['D2'] = info['name']
    ws['D2'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B3'] = info['date']
    ws['B3'].alignment = Alignment(vertical='center', horizontal='left')

    ws['D3'] = info.get('price', '')
    ws['D3'].alignment = Alignment(vertical='center', horizontal='left')

    ws['B4'] = info['desc']
    ws['B4'].alignment = Alignment(vertical='center', horizontal='left')

    step = resolveDepOwnerStepFromAudit(activity)
    ws['B5'] = getattr(step, 'desc', '无')
    ws['B5'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='hr', pos='owner')
    ws['B6'] = getattr(step, 'desc', '无')
    ws['B6'].alignment = Alignment(vertical='center', wrapText=True)

    step = resolveStepFromAudit(activity, dep='root', pos='ceo')
    ws['B8'] = getattr(step, 'desc', '无')
    ws['B8'].alignment = Alignment(vertical='center', wrapText=True)

    fix_merged_cells_border(ws, 'A2:D8')
    set_border(ws, 'A2:D8', 'medium')

    ws.protection.sheet = True
    ws.protection.set_password('zyhr2018')

    path = '/tmp/{}.xlsx'.format(str(uuid.uuid4()))
    wb.save(path)
    return path


def _export(activity):
    path, filename = None, None
    if activity.config.subtype == 'open_account':
        path = exportOpenAccountAuditDoc(activity)
        filename = '银行开户申请审批单.xlsx'
    elif re.match('cost', activity.config.subtype):
        path = exportCostAuditDoc(activity)
        filename = '费用报销审批单.xlsx'
    elif re.match('loan', activity.config.subtype):
        path = exportLoanAuditDoc(activity)
        filename = '借款审批单.xlsx'
    elif re.match('money', activity.config.subtype):
        path = exportMoneyAuditDoc(activity)
        filename = '用款审批单.xlsx'
    elif re.match('biz', activity.config.subtype):
        path = exportBizContractAuditDoc(activity)
        filename = '业务合同会签审批.xlsx'
    elif re.match('fn', activity.config.subtype):
        path = exportFnContractAuditDoc(activity)
        filename = '职能合同会签审批.xlsx'
    elif re.match('travel', activity.config.subtype):
        # travel
        path = exportTravelAuditDoc(activity)
        filename = '差旅费用报销审批单.xlsx'
    elif activity.config.subtype == 'yongren':
        # yongren
        path = exportYongrenAuditDoc(activity)
        filename = '用人需求审批单.xlsx'
    elif activity.config.subtype == 'qingjia':
        path = exportQingjiaAuditDoc(activity)
        filename = '请假申请审批单.xlsx'
    elif activity.config.subtype == 'chuchai':
        path = exportChuchaiAuditDoc(activity)
        filename = '出差申请审批单.xlsx'
    elif activity.config.subtype == 'kaoqin_yichang':
        path = exportKaoqinyichangAuditDoc(activity)
        filename = '考勤异常申请审批单.xlsx'
    elif activity.config.subtype == 'zizhi_shiyong':
        path = exportZizhishiyongAuditDoc(activity)
        filename = '资质使用申请审批单.xlsx'
    elif activity.config.subtype == 'yinjian_kezhi':
        path = exportYinjiankezhiAuditDoc(activity)
        filename = '印鉴刻制申请审批单.xlsx'
    elif activity.config.subtype == 'dangan_jiechu':
        path = exportDanganjiechuAuditDoc(activity)
        filename = '业务档案原件借出申请审批单.xlsx'
    elif activity.config.subtype == 'zichan_baofei':
        path = exportZichanbaofeiAuditDoc(activity)
        filename = '资产报废申请审批单.xlsx'
    else:
        pass

    return path, filename


@require_http_methods(['GET'])
def export(request, activityId):
    activity = AuditActivity.objects.get(pk=activityId)
    path, filename = _export(activity)
    return sendfile(request, path,
                    attachment=True,
                    attachment_filename=filename)


def zipdir(path, ziph):
    for root, dirs, files in os.walk(path):
        for file in files:
            ziph.write(os.path.join(root, file), arcname=file)


@require_http_methods(['GET'])
def batchExport(request):
    idx = request.GET.get('idx')
    idx = idx.split(',')

    dir = '/tmp/archive-' + str(uuid.uuid4())
    os.makedirs(dir)

    activities = AuditActivity.objects.filter(pk__in=idx)
    for activity in activities:
        path, filename = _export(activity)
        name = filename.split('.')[0]
        time_str = activity.created_at.strftime('%Y%m%d%H%M%S')
        copyfile(path, dir + '/{}-{}.xlsx'.format(name, time_str))

    path = '/tmp/审批单-{}.zip'.format(str(uuid.uuid4()))
    zipf = zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED)
    zipdir(dir, zipf)
    zipf.close()

    return sendfile(request, path,
                    attachment=True,
                    attachment_filename='审批单.zip')
